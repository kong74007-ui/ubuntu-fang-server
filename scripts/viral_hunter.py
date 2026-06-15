#!/usr/bin/env python3
"""
爆款视频识别 + detail补抓命令生成 — 两阶段抓取方法（方案B）

阶段1: search 模式全量扫(评论上限100/视频),快速覆盖
阶段2: 自动筛出爆款(点赞>1万 或 评论>1000),输出 detail 模式高量补抓命令(300条/视频)

数据合并: search + detail 两份 jsonl → 按 comment_id 去重 → 导出合并 Excel

用法：
  # 从 search_contents 筛爆款
  python scripts/viral_hunter.py data/douyin/jsonl/search_contents_*.jsonl

  # 直接输出可执行的补抓命令
  python scripts/viral_hunter.py data/douyin/jsonl/search_contents_*.jsonl --command

  # 合并 search + detail 两份数据 → 去重导出 xlsx
  python scripts/viral_hunter.py data/douyin/jsonl/ --merge --out merged.xlsx

实测效果（医美获客,59视频搜索→15爆款补抓）：
  评论 4002 → 4159(+2702 detail) → 6048(去重) ≈ 覆盖率翻倍
"""
import sys, os, json, glob, argparse


def load(path):
    rows = []
    for p in glob.glob(path):
        with open(p, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    rows.append(json.loads(line))
    return rows


def parse_count(v):
    """兼容 '104081' / '1.4万' 等格式"""
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if not s:
        return 0
    try:
        return int(s)
    except ValueError:
        if "万" in s:
            try:
                return int(float(s.replace("万", "")) * 10000)
            except ValueError:
                return 0
        return 0


def find_viral(works, like_thresh=10000, comment_thresh=1000):
    """筛选爆款视频（去重后按点赞降序）"""
    viral, seen = [], set()
    for w in works:
        aid = w.get("aweme_id", "")
        if aid in seen:
            continue
        seen.add(aid)
        likes = parse_count(w.get("liked_count", 0))
        cmts = parse_count(w.get("comment_count", 0))
        if likes >= like_thresh or cmts >= comment_thresh:
            viral.append((likes, cmts, w))
    viral.sort(key=lambda x: -x[0])
    return viral


def show_viral(viral):
    """打印爆款列表"""
    print(f"爆款视频: {len(viral)} 个 (点赞>{10000} 或 评论>{1000})\n")
    for i, (l, c, w) in enumerate(viral, 1):
        title = (w.get("title") or w.get("desc") or "")[:55]
        print(f"{i:2}. [{l}赞 / {c}评] {title}")
        print(f"    id: {w['aweme_id']}")
    return [w["aweme_id"] for _, _, w in viral]


def gen_command(ids, max_comments=300):
    """生成 detail 补抓命令"""
    print(f"\n=== detail 补抓命令（复制到终端执行）===\n")
    print(f"# 补抓 {len(ids)} 个爆款视频，每视频 {max_comments} 条评论")
    cmd = (
        f'uv run main.py --platform dy --lt qrcode --type detail '
        f'--specified_id "{",".join(ids)}" '
        f'--max_comments_count_singlenotes {max_comments}'
    )
    print(cmd)
    return cmd


def merge_and_export(data_dir, out_path):
    """合并 search + detail 数据，去重导出 xlsx"""
    data_dir = data_dir.rstrip("/\\")
    search_c = load(os.path.join(data_dir, "search_comments_*.jsonl"))
    search_v = load(os.path.join(data_dir, "search_contents_*.jsonl"))
    detail_c = load(os.path.join(data_dir, "detail_comments_*.jsonl"))
    detail_v = load(os.path.join(data_dir, "detail_contents_*.jsonl"))

    # 视频去重
    v_seen, videos = set(), []
    for d in search_v + detail_v:
        aid = d.get("aweme_id")
        if aid and aid not in v_seen:
            v_seen.add(aid)
            videos.append(d)

    # 评论去重(detail 优先覆盖)
    c_dict = {}
    for d in search_c:
        cid = d.get("comment_id")
        if cid:
            c_dict[cid] = d
    detail_from = 0
    for d in detail_c:
        cid = d.get("comment_id")
        if cid:
            c_dict[cid] = d
            detail_from += 1

    comments = list(c_dict.values())
    print(f"合并结果: 视频 {len(videos)} / 评论 {len(comments)} (detail来源 {detail_from})")

    # 导出 xlsx
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HF = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    HFT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    BF = Font(name="微软雅黑", size=10)
    B = Border(left=Side(style="thin"), right=Side(style="thin"),
               top=Side(style="thin"), bottom=Side(style="thin"))

    wb = Workbook()
    ws = wb.active; ws.title = "评论"
    headers = ["序号","评论ID","视频ID","评论内容","点赞","IP属地","用户ID","抖音号","昵称","用户主页","来源"]
    ws.append(headers)
    for ci in range(1, len(headers)+1):
        c = ws.cell(row=1, column=ci); c.fill = HF; c.font = HFT; c.border = B
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    detail_awemes = {d["aweme_id"] for d in detail_v}
    for i, cm in enumerate(comments, 1):
        sec = cm.get("sec_uid") or ""
        src = "detail" if cm.get("aweme_id") in detail_awemes else "search"
        vals = [i, cm.get("comment_id"), cm.get("aweme_id"), cm.get("content"),
                cm.get("like_count"), cm.get("ip_location"), cm.get("user_id"),
                cm.get("user_unique_id"), cm.get("nickname"),
                f"https://www.douyin.com/user/{sec}" if sec else "", src]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=i+1, column=ci, value=v); c.font = BF; c.border = B
    ws.column_dimensions["D"].width = 55; ws.column_dimensions["J"].width = 42
    wb.save(out_path)
    print(f"导出完成: {out_path}")


def main():
    ap = argparse.ArgumentParser(description="爆款视频识别 + detail补抓命令生成")
    ap.add_argument("path", help="search_contents jsonl路径或含search/detail子文件的目录")
    ap.add_argument("--command", action="store_true", help="直接输出可执行的 detail 命令")
    ap.add_argument("--merge", action="store_true", help="合并 search+detail 并导出xlsx")
    ap.add_argument("--out", default="merged.xlsx", help="合并导出路径(默认 merged.xlsx)")
    ap.add_argument("--detail-comments", type=int, default=300, help="detail 模式评论上限(默认300)")
    ap.add_argument("--like-thresh", type=int, default=10000, help="爆款点赞阈值(默认10000)")
    ap.add_argument("--comment-thresh", type=int, default=1000, help="爆款评论数阈值(默认1000)")
    args = ap.parse_args()

    if args.merge:
        merge_and_export(args.path, args.out)
        return

    works = load(args.path)
    print(f"视频总数(去重): {len(works)}")
    viral = find_viral(works, args.like_thresh, args.comment_thresh)
    ids = show_viral(viral)

    if args.command and ids:
        gen_command(ids, args.detail_comments)


if __name__ == "__main__":
    main()
