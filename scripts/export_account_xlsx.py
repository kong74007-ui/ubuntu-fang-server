#!/usr/bin/env python3
"""
账号 xlsx 导出器 —— 把 MediaCrawler creator 模式的输出导成「一个账号一个 xlsx」。
含 3 个 sheet：① 账号画像 ② 全部作品(18列) ③ 评论。

用法：python export_account_xlsx.py <jsonl_dir> <输出.xlsx>
"""
import sys
import os
import glob
import json
from datetime import datetime, timezone, timedelta
import openpyxl

CST = timezone(timedelta(hours=8))
WORK_HEADERS = ["序号", "视频ID", "标题", "描述", "发布时间", "用户ID", "抖音号", "昵称",
                "点赞", "收藏", "评论数", "分享", "IP属地", "视频链接", "用户主页",
                "封面", "视频下载", "来源关键词"]
CMT_HEADERS = ["序号", "评论内容", "昵称", "抖音号", "属地", "点赞", "所属视频ID"]


def load(pattern):
    rows = []
    for p in glob.glob(pattern):
        with open(p) as f:
            for line in f:
                line = line.strip()
                if line:
                    rows.append(json.loads(line))
    return rows


def ftime(ts):
    try:
        return datetime.fromtimestamp(int(ts), CST).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ""


def main():
    jdir, out = sys.argv[1], sys.argv[2]
    creators = load(os.path.join(jdir, "creator_creators_*.jsonl"))
    works = load(os.path.join(jdir, "creator_contents_*.jsonl"))
    comments = load(os.path.join(jdir, "creator_comments_*.jsonl"))

    sec = works[0].get("sec_uid") if works else ""
    wb = openpyxl.Workbook()

    # sheet1 画像
    ws = wb.active
    ws.title = "账号画像"
    c = creators[0] if creators else {}
    ip = (c.get("ip_location") or "").replace("IP属地：", "")
    ws.append(["项目", "数据"])
    for k, v in [("昵称", c.get("nickname")), ("抖音号", works[0].get("user_unique_id") if works else ""),
                 ("粉丝", c.get("fans")), ("获赞", c.get("interaction")),
                 ("作品数", c.get("videos_count")), ("关注", c.get("follows")),
                 ("IP属地", ip), ("简介", c.get("desc") or "无"),
                 ("主页", f"https://www.douyin.com/user/{sec}" if sec else "")]:
        ws.append([k, v])
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 60

    # sheet2 作品
    ws2 = wb.create_sheet("全部作品")
    ws2.append(WORK_HEADERS)
    for i, d in enumerate(works, 1):
        s = d.get("sec_uid") or ""
        ws2.append([i, d.get("aweme_id"), d.get("title"), d.get("desc"), ftime(d.get("create_time")),
                    d.get("user_id"), d.get("user_unique_id"), d.get("nickname"),
                    d.get("liked_count"), d.get("collected_count"), d.get("comment_count"),
                    d.get("share_count"), d.get("ip_location"), d.get("aweme_url"),
                    f"https://www.douyin.com/user/{s}" if s else "",
                    d.get("cover_url"), d.get("video_download_url"), d.get("source_keyword")])
    for col, w in {"C": 40, "D": 40, "N": 34, "O": 34}.items():
        ws2.column_dimensions[col].width = w

    # sheet3 评论
    ws3 = wb.create_sheet("评论")
    ws3.append(CMT_HEADERS)
    for i, c in enumerate(comments, 1):
        ws3.append([i, c.get("content"), c.get("nickname"), c.get("user_unique_id"),
                    c.get("ip_location"), c.get("like_count"), c.get("aweme_id")])
    ws3.column_dimensions["B"].width = 50

    wb.save(out)
    print(f"✅ 账号「{c.get('nickname')}」→ {out}")
    print(f"   画像: 粉丝{c.get('fans')}/获赞{c.get('interaction')}/作品{c.get('videos_count')} | "
          f"作品表 {len(works)} 条 | 评论 {len(comments)} 条")


if __name__ == "__main__":
    main()
